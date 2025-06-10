from rest_framework import viewsets
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from ..models import Result
from ..serializers.result_serializers import ResultListSerializers, ResultRetrieveSerializers, ResultWriteSerializers
from ..utilities.importbase import *
import csv
import io
import json
from rest_framework.permissions import IsAuthenticated
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework.response import Response
from django.db import transaction
from openpyxl import load_workbook  # for .xlsx/.xls support
from mainproj.permissions import DynamicModelPermission
from ..models import File, Result
from ..serializers.result_serializers import (
    ResultListSerializers,
    ResultRetrieveSerializers,
    ResultWriteSerializers
)
from ..utilities.permissions import resultmanagementPermission
from ..utilities.pagination import MyPageNumberPagination

CACHE_TTL = 60 * 60 * 24 * 5   # cache for 5 days 


class resultViewsets(viewsets.ModelViewSet):
    """
    Exposes /api/results/ for standard list/retrieve/create/update/delete,
    plus a custom POST /api/results/bulk-import/ for batch‐importing a CSV/Excel file.
    """
    queryset = Result.objects.all().order_by("symbol_no")
    serializer_class = ResultListSerializers
    # permission_classes = [DynamicModelPermission]
    pagination_class = MyPageNumberPagination
    filter_backends = [SearchFilter, DjangoFilterBackend, OrderingFilter]
    filterset_fields = {
        'id': ['exact'],
        'symbol_no': ['exact'],
        'file': ['exact'],
        'created_date': ['exact', 'gte', 'lte'],
        'updated_date': ['exact', 'gte', 'lte'],
    }


    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return ResultWriteSerializers
        elif self.action == "retrieve":
            return ResultRetrieveSerializers
        return super().get_serializer_class()
    
    @method_decorator(cache_page(CACHE_TTL), name="list")
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @method_decorator(cache_page(CACHE_TTL), name="retrieve")
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-import",
        permission_classes=[IsAuthenticated],  # enable if needed
    )
    def bulk_import(self, request, *args, **kwargs):
        """
        Expects JSON body:
        {
          "file_id": <existing File PK>,
          "name":    "<display name to store on File.name>",
          "mapping": {
            // Required:
            "symbol_no":    "<column_name_for_symbol_no>",
            "dateofbirth":  "<column_name_for_dateofbirth>",
            "cgpa":         "<column_name_for_cgpa>",
            "remarks":      "<column_name_for_remarks>",
            // Optional:
            "student_name": "<column_name_for_student_name>"
          }
        }

        Behavior:
        1) Look up File by ID.
        2) If any Result rows exist for that File, delete them. Otherwise skip deletion.
        3) Save top‐level "name" onto File.name.
        4) Save mapping JSON onto File.mapped_json.
        5) Stream through File.file (CSV or Excel), bulk-create new Result rows:
             - If "student_name" was in mapping, use that column; otherwise default to name_value.
        """
        # ─── 1) Read top‐level inputs ─────────────────────────────────────
        file_id = request.data.get("file_id")
        name_value = request.data.get("name", "").strip()
        mapping_json = request.data.get("mapping")

        # Validate file_id
        if file_id is None:
            return Response(
                {"error": "Missing 'file_id'. Please send an existing File ID."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            file_obj = File.objects.get(pk=file_id)
        except File.DoesNotExist:
            return Response(
                {"error": f"File with ID={file_id} not found."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate name (to store on File.name)
        if not name_value:
            return Response(
                {"error": "Missing 'name'. Please send a non-empty name string."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parse mapping JSON
        if mapping_json is None:
            return Response(
                {"error": "Missing 'mapping'. Please send JSON under key 'mapping'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if isinstance(mapping_json, str):
            try:
                mapping = json.loads(mapping_json)
            except json.JSONDecodeError as e:
                return Response(
                    {"error": f"Invalid JSON in 'mapping': {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        elif isinstance(mapping_json, dict):
            mapping = mapping_json
        else:
            return Response(
                {"error": "'mapping' must be a JSON string or JSON object."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Ensure the four required keys are present; "student_name" is optional
        required_keys = {"symbol_no", "dateofbirth", "cgpa", "remarks"}
        if not required_keys.issubset(set(mapping.keys())):
            return Response(
                {
                    "error": (
                        "'mapping' must contain at least these keys: "
                        "symbol_no, dateofbirth, cgpa, remarks"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ─── 2) Delete existing Results for this File only if they exist ─
        if Result.objects.filter(file=file_obj).exists():
            Result.objects.filter(file=file_obj).delete()

        # ─── 3) Save name_value onto File.name ───────────────────────────
        file_obj.name = name_value
        file_obj.save(update_fields=["name"])

        # ─── 4) Save mapping JSON onto File.mapped_json ─────────────────
        file_obj.mapped_json = mapping
        file_obj.save(update_fields=["mapped_json"])

        # ─── 5) Determine file type by extension of the stored file ─────
        filename_lower = file_obj.file.name.lower()
        if filename_lower.endswith(".csv"):
            file_type = "csv"
        elif filename_lower.endswith((".xlsx", ".xls")):
            file_type = "excel"
        else:
            return Response(
                {"error": "Unsupported file format. Must be .csv, .xlsx, or .xls."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        BATCH_SIZE = 5000
        instances_to_create = []

        def normalize_header(hdr: str) -> str:
            return str(hdr).strip()

        try:
            with transaction.atomic():
                if file_type == "csv":
                    # ─── CSV path: open in binary then wrap for text ───────────
                    file_obj.file.open("rb")
                    text_stream = io.TextIOWrapper(
                        file_obj.file, encoding="utf-8", newline=""
                    )
                    reader = csv.DictReader(text_stream)

                    # Check header row
                    csv_headers = [normalize_header(h) for h in (reader.fieldnames or [])]

                    # Verify each required column exists
                    for col_key in required_keys:
                        col_name = mapping[col_key]
                        if normalize_header(col_name) not in csv_headers:
                            raise KeyError(f"Column '{col_name}' not found in CSV header.")

                    # If "student_name" was provided, verify it too
                    has_name_column = False
                    if "student_name" in mapping:
                        name_col = mapping["student_name"]
                        if normalize_header(name_col) not in csv_headers:
                            raise KeyError(f"Column '{name_col}' not found in CSV header.")
                        has_name_column = True

                    # Stream each row
                    for row in reader:
                        try:
                            symbol_no    = row[normalize_header(mapping["symbol_no"])].strip()
                            dateofbirth  = row[normalize_header(mapping["dateofbirth"])].strip()
                            cgpa         = row[normalize_header(mapping["cgpa"])].strip()
                            remarks      = row[normalize_header(mapping["remarks"])].strip()

                            if has_name_column:
                                student_name = row[normalize_header(mapping["student_name"])].strip()
                            else:
                                student_name = name_value

                            instances_to_create.append(
                                Result(
                                    student_name=student_name,
                                    file=file_obj,
                                    symbol_no=symbol_no,
                                    dateofbirth=dateofbirth,
                                    cgpa=cgpa,
                                    remarks=remarks,
                                )
                            )
                        except KeyError:
                            # Skip any row missing required columns
                            continue
                        except Exception:
                            # Skip any other malformed row
                            continue

                        # Bulk-insert in batches
                        if len(instances_to_create) >= BATCH_SIZE:
                            Result.objects.bulk_create(instances_to_create)
                            instances_to_create = []

                    # Flush any remaining rows
                    if instances_to_create:
                        Result.objects.bulk_create(instances_to_create)
                        instances_to_create = []

                    file_obj.file.close()

                else:
                    # ─── Excel path: open in binary and pass to openpyxl ────
                    file_obj.file.open("rb")
                    wb = load_workbook(
                        filename=file_obj.file, read_only=True, data_only=True
                    )
                    sheet = wb.active
                    excel_iter = sheet.iter_rows(values_only=True)

                    header_row = next(excel_iter, None)
                    if not header_row:
                        raise KeyError("Excel file appears empty (no header).")

                    # Build header_map: { normalized_header: index }
                    header_map = {
                        normalize_header(cell_val): idx
                        for idx, cell_val in enumerate(header_row)
                        if cell_val is not None
                    }

                    # Verify each required column exists
                    for col_key in required_keys:
                        col_name = mapping[col_key]
                        if normalize_header(col_name) not in header_map:
                            raise KeyError(f"Column '{col_name}' not found in Excel headers.")

                    # If "student_name" provided, verify it too
                    has_name_column = False
                    if "student_name" in mapping:
                        name_col = mapping["student_name"]
                        if normalize_header(name_col) not in header_map:
                            raise KeyError(f"Column '{name_col}' not found in Excel headers.")
                        has_name_column = True

                    # Stream each subsequent row
                    for row_tuple in excel_iter:
                        try:
                            sn_idx  = header_map[normalize_header(mapping["symbol_no"])]
                            dob_idx = header_map[normalize_header(mapping["dateofbirth"])]
                            cg_idx  = header_map[normalize_header(mapping["cgpa"])]
                            rm_idx  = header_map[normalize_header(mapping["remarks"])]

                            if has_name_column:
                                name_idx = header_map[normalize_header(mapping["student_name"])]

                            def get_cell(r, i):
                                return "" if (i >= len(r) or r[i] is None) else str(r[i]).strip()

                            if has_name_column:
                                student_name = get_cell(row_tuple, name_idx)
                            else:
                                student_name = name_value

                            instances_to_create.append(
                                Result(
                                    student_name=student_name,
                                    file=file_obj,
                                    symbol_no=get_cell(row_tuple, sn_idx),
                                    dateofbirth=get_cell(row_tuple, dob_idx),
                                    cgpa=get_cell(row_tuple, cg_idx),
                                    remarks=get_cell(row_tuple, rm_idx),
                                )
                            )
                        except KeyError:
                            continue
                        except Exception:
                            continue

                        if len(instances_to_create) >= BATCH_SIZE:
                            Result.objects.bulk_create(instances_to_create)
                            instances_to_create = []

                    if instances_to_create:
                        Result.objects.bulk_create(instances_to_create)
                        instances_to_create = []

                    wb.close()
                    file_obj.file.close()

        except KeyError as ke:
            return Response(
                {"error": f"Mapping error: {str(ke)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            return Response(
                {"error": f"Server error while importing: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {"message": "File processed and data saved to database."},
            status=status.HTTP_200_OK,
        )